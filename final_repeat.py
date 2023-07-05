''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
  * @file : excel數據轉換_載具使用數量及使用比率
  * @brief : 需根據最終檔案學校名的排序逐一設定需讀取檔案的名稱為{流水序}

  * 檔案備註 : DONE
  
  * @author : 黃奕維
  * @date : 2023/7/5
'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
import openpyxl
from openpyxl.styles import Font 
import os


FinalFileName="鳳林鎮"
REPEAT=9      #重複次數
R=2             #開始寫入列數
C=4     #開始寫入行數


for number in range(REPEAT ):

    os.chdir(r"C:\Users\USER\Desktop\python excel\file")# Excel檔案的絕對路徑
    wb = openpyxl.load_workbook( 'DeviceQuantityExcel ('+str(number)+')'+'.xlsx')#開啟要讀取的Excel檔案
    #wb = openpyxl.load_workbook(str(number+1)+'.xlsx')#開啟要讀取的Excel檔案
    wb_s1= wb.worksheets[0]#讀取檔案裡的第一個工作表
    wb_s1_maxcolumn=wb_s1.max_column#檔案裡的行數
    wb_s1_maxrow=wb_s1.max_row#檔案裡的列數

    done = openpyxl.load_workbook(FinalFileName+'.xlsx')#開啟要複寫的Excel檔案
    done_s1=done.worksheets[0]#讀取檔案裡的第一個工作表
    done_s1_maxcolumn=done_s1.max_column#檔案裡的行數
    done_s1_maxrow=done_s1.max_row#檔案裡的列數
    data=[]#存放讀取到的資料
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''      #read
    for c in range(2,wb_s1_maxcolumn+1):  #再讀行
        for r in range(2,wb_s1_maxrow+1):  #先讀列
            value=wb_s1.cell(r,c).value
            #print(wb_s1.cell(r,c).value)#TEST
            data.append(value)#存入資料
    print("DATA",str(data))#TEST
    a=0#TEST
    #print(done_s1_maxcolumn)#TEST
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''       #write
    for c in range(C,31):   #讀行
        #a=a+1#TEST
        #print(a)#TEST
        done_s1.cell(R,c).value=data[c-C]#將資料寫入
        #print("c",str(c))
        #print("c-C",str(c-C))
        #print(data[c-C])#TEST
        
    
    R=R+1#往下寫一行
    done.save(FinalFileName+'.xlsx')#儲存複寫檔
    
    print(number,"  Finish")
print("DONEEEEEEEEEEEEEEE!!!!!!!!")







"""
sheet.cell(rowNum, 2).font = Font(color='FF0000')
"""


